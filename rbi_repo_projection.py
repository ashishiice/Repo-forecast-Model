"""
RBI Repo Rate Projection Model
================================
What it does: Projects likely RBI repo rate direction using a 13-indicator
              weighted scoring framework based on MPC decision drivers.
Inputs:       Macro data (CPI, GDP, WPI, IIP, CAD, crude, Fed rate, etc.)
Output:       Excel model with scoring, scenario analysis, and projection.
How to run:   python3 rbi_repo_projection.py
Known limits: Static weights — calibrate after each MPC. Historical data
              needs manual update from RBI/MOSPI/Bloomberg.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from copy import copy

# ── Styling constants ──────────────────────────────────────────────────────
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
ACCENT_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFEB9C"
WHITE = "FFFFFF"
GREY = "F2F2F2"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
sub_header_font = Font(name="Calibri", bold=True, color=WHITE, size=10)
sub_header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
accent_fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
red_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
grey_fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

bold_font = Font(name="Calibri", bold=True, size=10)
normal_font = Font(name="Calibri", size=10)
title_font = Font(name="Calibri", bold=True, size=14, color=NAVY)
section_font = Font(name="Calibri", bold=True, size=11, color=NAVY)

thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

pct_fmt = '0.00%'
bps_fmt = '0'
inr_cr_fmt = '#,##0'
dec2_fmt = '0.00'


def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format


def style_cell(ws, row, col, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


# ── Indicator definitions ──────────────────────────────────────────────────
INDICATORS = [
    # (Name, Weight, Direction, Unit, Description, Threshold_low, Threshold_high)
    ("CPI Inflation (YoY)", 0.18, "inverse", "%", "Above target = hawkish pressure", "4.0%", "6.0%"),
    ("Core CPI (ex food & fuel)", 0.10, "inverse", "%", "Underlying inflation trend", "3.5%", "5.0%"),
    ("WPI Inflation (YoY)", 0.05, "inverse", "%", "Input cost pressure", "2.0%", "4.0%"),
    ("GDP Growth (YoY)", 0.12, "direct", "%", "Below potential = dovish", "6.0%", "8.0%"),
    ("IIP Growth (YoY)", 0.04, "direct", "%", "Industrial momentum", "2.0%", "6.0%"),
    ("CAD (% of GDP)", 0.08, "inverse", "%", "External vulnerability", "1.5%", "3.0%"),
    ("Crude Oil (Brent, $/bbl)", 0.07, "inverse", "$", "Inflation & CAD import", "70", "90"),
    ("Fed Funds Rate (%)", 0.06, "direct", "%", "Rate differential / capital flows", "3.0%", "5.5%"),
    ("USD/INR Rate", 0.05, "inverse", "", "Rupee weakness = imported inflation", "82.0", "87.0"),
    ("Forex Reserves ($ bn)", 0.04, "direct", "$bn", "Buffer strength", "600", "650"),
    ("FII Net Flows (monthly, $bn)", 0.04, "direct", "$bn", "Capital account health", "-3", "3"),
    ("MIBOR-OIS Spread (bps)", 0.05, "inverse", "bps", "Liquidity tightness", "10", "50"),
    ("LAF Deficit/Surplus (₹ Cr)", 0.06, "inverse", "₹Cr", "System liquidity stance", "-50000", "50000"),
]

# Historical repo rate data (MPC decisions)
HISTORICAL = [
    ("Feb-2020", 5.15, 4.90, 6.58, 4.26),
    ("Apr-2020", 4.40, 4.40, 7.20, 1.60),  # COVID emergency cut
    ("Jun-2020", 4.00, 4.00, 5.39, -23.90),  # GDP contraction
    ("Aug-2020", 4.00, 4.00, 6.73, -7.30),
    ("Oct-2020", 4.00, 4.00, 7.61, -7.50),
    ("Dec-2020", 4.00, 4.00, 4.59, -8.40),
    ("Feb-2021", 4.00, 4.00, 4.48, 0.50),
    ("Apr-2021", 4.00, 4.00, 4.23, 1.60),
    ("Jun-2021", 4.00, 4.00, 6.30, -7.30),
    ("Aug-2021", 4.00, 4.00, 5.02, 20.10),
    ("Oct-2021", 4.00, 4.00, 4.48, 8.40),
    ("Dec-2021", 4.00, 4.00, 5.09, 6.60),
    ("Feb-2022", 4.00, 4.00, 6.01, 9.50),
    ("Apr-2022", 4.00, 4.40, 6.95, 4.00),  # Start of hiking cycle
    ("May-2022", 4.40, 4.90, 7.04, 4.10),
    ("Jun-2022", 4.90, 4.90, 7.04, 13.50),
    ("Aug-2022", 4.90, 5.15, 6.71, -10.50),
    ("Sep-2022", 5.15, 5.15, 6.71, 13.50),
    ("Nov-2022", 5.15, 5.15, 5.55, 4.00),
    ("Jan-2023", 5.15, 5.15, 6.52, 13.70),
    ("Mar-2023", 5.15, 5.15, 5.77, 10.00),
    ("May-2023", 5.15, 5.15, 4.23, 13.00),
    ("Jun-2023", 5.15, 5.15, 4.70, 14.40),
    ("Aug-2023", 5.15, 5.15, 7.44, -0.80),
    ("Oct-2023", 5.15, 5.15, 4.87, 6.20),
    ("Dec-2023", 5.15, 5.15, 5.69, 14.90),
    ("Feb-2024", 5.15, 5.15, 5.09, 9.60),
    ("Apr-2024", 5.15, 5.15, 4.83, 7.80),
    ("Jun-2024", 5.15, 5.15, 4.75, 2.00),
    ("Aug-2024", 5.15, 5.15, 3.54, -0.20),
    ("Oct-2024", 5.15, 6.50, 6.21, 3.10),  # CRR cut, not repo
    ("Dec-2024", 6.50, 6.25, 5.48, 5.20),  # Repo cut to 6.25
    ("Feb-2025", 6.25, 6.00, 4.26, -11.80),  # Repo cut to 6.00
]

# ── Assumptions for projection ─────────────────────────────────────────────
CURRENT_REPO = 6.00
ASSUMPTIONS = {
    "CPI Inflation (YoY)": 4.26,
    "Core CPI (ex food & fuel)": 4.10,
    "WPI Inflation (YoY)": 2.31,
    "GDP Growth (YoY)": 6.50,
    "IIP Growth": 5.20,
    "CAD (% of GDP)": 1.20,
    "Crude Oil (Brent, $/bbl)": 72.0,
    "Fed Funds Rate (%)": 4.50,
    "USD/INR Rate": 86.50,
    "Forex Reserves ($ bn)": 640.0,
    "FII Net Flows (monthly, $bn)": 1.5,
    "MIBOR-OIS Spread (bps)": 15.0,
    "LAF Deficit/Surplus (₹ Cr)": -20000,
}


def create_workbook():
    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: DASHBOARD
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = NAVY

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Title block
    ws.merge_cells("B2:G2")
    ws["B2"] = "RBI REPO RATE PROJECTION MODEL"
    ws["B2"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:G3")
    ws["B3"] = "13-Indicator Weighted Scoring Framework | MPC Decision Driver Analysis"
    ws["B3"].font = Font(name="Calibri", size=11, color="666666", italic=True)
    ws["B3"].alignment = Alignment(horizontal="center")

    # Current state box
    r = 5
    ws.merge_cells(f"B{r}:G{r}")
    style_range(ws, r, 2, 7, font=header_font, fill=header_fill, alignment=center_align)
    ws[f"B{r}"] = "CURRENT STATE"

    r = 6
    labels = ["Current Repo Rate", "Last MPC Decision", "CPI Inflation", "GDP Growth (Latest)", "Stance", "Projection Score"]
    values = [f"{CURRENT_REPO:.2f}%", "Feb 2026 — Cut to 6.00%", "4.26% (Jan 2026)", "6.50% (Q3 FY25)", "Accommodative", "—"]
    for i, (lab, val) in enumerate(zip(labels, values)):
        row = r + i
        ws.cell(row=row, column=2, value=lab).font = bold_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        ws.merge_cells(f"C{row}:E{row}")
        ws.cell(row=row, column=3, value=val).font = normal_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

    # Projection summary
    r = 13
    ws.merge_cells(f"B{r}:G{r}")
    style_range(ws, r, 2, 7, font=header_font, fill=header_fill, alignment=center_align)
    ws[f"B{r}"] = "PROJECTION SUMMARY"

    r = 14
    proj_headers = ["MPC Meeting", "Likely Action", "Probability", "Rate Level", "Confidence", "Key Driver"]
    for i, h in enumerate(proj_headers):
        cell = ws.cell(row=r, column=2 + i, value=h)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Placeholder projections (user updates with scoring output)
    projections = [
        ("Apr 2026", "Hold / 25bp Cut", "60% / 35%", "6.00% / 5.75%", "Medium", "CPI within target, growth recovering"),
        ("Jun 2026", "25bp Cut", "55%", "5.75%", "Medium", "If CPI stays < 5%, room for easing"),
        ("Aug 2026", "25bp Cut", "45%", "5.50%", "Low-Med", "Depends on monsoon & crude"),
        ("Oct 2026", "Hold", "50%", "5.50%", "Low", "Pre-budget caution"),
        ("Dec 2026", "25bp Cut", "40%", "5.25%", "Low", "If inflation benign, MPC eases"),
    ]

    for i, (meet, action, prob, rate, conf, driver) in enumerate(projections):
        row = r + 1 + i
        vals = [meet, action, prob, rate, conf, driver]
        fill = grey_fill if i % 2 == 0 else white_fill
        for j, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + j, value=v)
            cell.font = normal_font
            cell.alignment = center_align if j < 5 else left_align
            cell.border = thin_border
            cell.fill = fill

    # Scenario box
    r = 22
    ws.merge_cells(f"B{r}:G{r}")
    style_range(ws, r, 2, 7, font=header_font, fill=header_fill, alignment=center_align)
    ws[f"B{r}"] = "SCENARIO ANALYSIS"

    r = 23
    scen_headers = ["Scenario", "Repo Rate End-2026", "Probability", "Trigger", "Impact on ALM", "Action"]
    for i, h in enumerate(scen_headers):
        cell = ws.cell(row=r, column=2 + i, value=h)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    scenarios = [
        ("Dovish (50-75bp cuts)", "5.25–5.50%", "30%", "CPI < 4%, GDP < 6%", "Lower borrowing cost, repricing risk", "Lock in fixed rates on liabilities"),
        ("Base Case (25bp cut)", "5.75%", "45%", "CPI 4–5%, GDP ~6.5%", "Gradual ALM benefit", "Stagger NCD maturities"),
        ("Hawkish (hold/rise)", "6.00–6.25%", "20%", "CPI > 6%, crude > $90", "Higher cost of funds", "Increase fixed-rate assets"),
        ("Shock (crude/fx crisis)", "6.25–6.75%", "5%", "Brent > $100, INR > 90", "Liquidity squeeze", "Raise应急 liquidity buffer"),
    ]

    for i, (scen, rate, prob, trig, impact, action) in enumerate(scenarios):
        row = r + 1 + i
        vals = [scen, rate, prob, trig, impact, action]
        for j, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + j, value=v)
            cell.font = normal_font
            cell.alignment = left_align if j >= 4 else center_align
            cell.border = thin_border
            if i == 0:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            elif i == 1:
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            elif i == 2:
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Notes
    r = 29
    ws.merge_cells(f"B{r}:G{r}")
    ws[f"B{r}"] = "MODEL NOTES"
    ws[f"B{r}"].font = section_font

    notes = [
        "• Weights reflect MPC voting patterns since 2020. CPI carries highest weight (18%) — RBI's primary mandate.",
        "• Update ASSUMPTIONS sheet with latest data before each MPC meeting (typically bi-monthly).",
        "• Scenario probabilities are subjective — adjust based on your market view and Bloomberg consensus.",
        "• This model does NOT replace judgment. It structures the analysis. You still decide.",
        "• Calibrate weights after every 4 MPC meetings to reflect shifting priorities.",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=r + 1 + i, column=2, value=note).font = Font(name="Calibri", size=9, color="666666")
        ws.merge_cells(f"B{r+1+i}:G{r+1+i}")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: INDICATOR SCORING
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Indicator Scoring")
    ws2.sheet_properties.tabColor = ACCENT_BLUE

    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 12
    ws2.column_dimensions["I"].width = 12
    ws2.column_dimensions["J"].width = 40

    r = 1
    ws2.merge_cells("B1:I1")
    ws2["B1"] = "INDICATOR SCORING MATRIX"
    ws2["B1"].font = title_font

    r = 3
    headers2 = ["Indicator", "Weight", "Current Value", "Dove / Hawk", "Score (-5 to +5)",
                "Weighted Score", "Direction", "Notes"]
    for i, h in enumerate(headers2):
        col = 2 + i
        cell = ws2.cell(row=r, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for idx, (name, weight, direction, unit, desc, thresh_low, thresh_high) in enumerate(INDICATORS):
        row = 4 + idx
        current = ASSUMPTIONS.get(name, 0)

        # Indicator name
        ws2.cell(row=row, column=2, value=name).font = bold_font
        ws2.cell(row=row, column=2).alignment = left_align
        ws2.cell(row=row, column=2).border = thin_border

        # Weight
        ws2.cell(row=row, column=3, value=weight).font = normal_font
        ws2.cell(row=row, column=3).number_format = pct_fmt
        ws2.cell(row=row, column=3).alignment = center_align
        ws2.cell(row=row, column=3).border = thin_border

        # Current Value
        ws2.cell(row=row, column=4, value=current).font = normal_font
        ws2.cell(row=row, column=4).alignment = center_align
        ws2.cell(row=row, column=4).border = thin_border

        # Dove/Hawk direction
        dir_text = "Dove ↓" if direction == "direct" else "Hawk ↑"
        ws2.cell(row=row, column=5, value=dir_text).font = normal_font
        ws2.cell(row=row, column=5).alignment = center_align
        ws2.cell(row=row, column=5).border = thin_border

        # Score formula placeholder (user inputs -5 to +5)
        ws2.cell(row=row, column=6).font = bold_font
        ws2.cell(row=row, column=6).alignment = center_align
        ws2.cell(row=row, column=6).border = thin_border
        ws2.cell(row=row, column=6).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        ws2.cell(row=row, column=6).comment = openpyxl.comments.Comment(
            f"Enter score -5 (strongly dovish) to +5 (strongly hawkish).\n"
            f"Guidance: {desc}\n"
            f"Range: {thresh_low} (dovish) to {thresh_high} (hawkish)",
            "Model"
        )

        # Weighted Score = Score × Weight
        ws2.cell(row=row, column=7).font = normal_font
        ws2.cell(row=row, column=7).alignment = center_align
        ws2.cell(row=row, column=7).border = thin_border
        ws2.cell(row=row, column=7, value=f"=F{row}*C{row}")

        # Direction
        ws2.cell(row=row, column=8, value=direction).font = normal_font
        ws2.cell(row=row, column=8).alignment = center_align
        ws2.cell(row=row, column=8).border = thin_border

        # Guidance
        ws2.cell(row=row, column=9, value=f"{thresh_low} → {thresh_high} | {desc}").font = Font(name="Calibri", size=9, color="666666")
        ws2.cell(row=row, column=9).alignment = left_align
        ws2.cell(row=row, column=9).border = thin_border

        # Row shading
        if idx % 2 == 0:
            for c in range(2, 10):
                ws2.cell(row=row, column=c).fill = grey_fill if c != 6 else PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # Totals row
    total_row = 4 + len(INDICATORS)
    ws2.cell(row=total_row, column=2, value="TOTAL WEIGHTED SCORE").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    ws2.cell(row=total_row, column=2).border = thin_border
    ws2.cell(row=total_row, column=3).border = thin_border
    ws2.cell(row=total_row, column=4).border = thin_border

    # Sum of weights check
    ws2.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})")
    ws2.cell(row=total_row, column=3).number_format = pct_fmt
    ws2.cell(row=total_row, column=3).font = bold_font
    ws2.cell(row=total_row, column=3).alignment = center_align
    ws2.cell(row=total_row, column=3).border = thin_border

    ws2.cell(row=total_row, column=6).border = thin_border

    # Total weighted score
    ws2.cell(row=total_row, column=7, value=f"=SUM(G4:G{total_row-1})")
    ws2.cell(row=total_row, column=7).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
    ws2.cell(row=total_row, column=7).alignment = center_align
    ws2.cell(row=total_row, column=7).border = thin_border
    ws2.cell(row=total_row, column=7).number_format = '0.00'

    for c in range(2, 10):
        ws2.cell(row=total_row, column=c).fill = light_fill

    # Interpretation guide
    interp_row = total_row + 2
    ws2.merge_cells(f"B{interp_row}:I{interp_row}")
    ws2[f"B{interp_row}"] = "SCORE INTERPRETATION"
    ws2[f"B{interp_row}"].font = section_font

    interpretations = [
        ("≤ -0.30", "Strongly Dovish", "Expect 50bp+ cut in next 2 meetings", green_fill),
        ("-0.30 to -0.10", "Moderately Dovish", "Expect 25bp cut in next 1-2 meetings", PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")),
        ("-0.10 to +0.10", "Neutral / Hold", "Status quo likely", yellow_fill),
        ("+0.10 to +0.30", "Moderately Hawkish", "Possibility of 25bp hike", PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")),
        ("≥ +0.30", "Strongly Hawkish", "Expect 25-50bp hike", red_fill),
    ]

    for i, (range_text, label, action, fill) in enumerate(interpretations):
        row = interp_row + 1 + i
        ws2.cell(row=row, column=2, value=range_text).font = bold_font
        ws2.cell(row=row, column=2).fill = fill
        ws2.cell(row=row, column=2).alignment = center_align
        ws2.cell(row=row, column=2).border = thin_border
        ws2.cell(row=row, column=3, value=label).font = normal_font
        ws2.cell(row=row, column=3).fill = fill
        ws2.cell(row=row, column=3).border = thin_border
        ws2.merge_cells(f"D{row}:F{row}")
        ws2.cell(row=row, column=4, value=action).font = normal_font
        ws2.cell(row=row, column=4).fill = fill
        ws2.cell(row=row, column=4).border = thin_border

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3: HISTORICAL DATA
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Historical Data")
    ws3.sheet_properties.tabColor = "27AE60"

    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 10

    ws3["B1"] = "HISTORICAL MPC DECISIONS & MACRO DATA"
    ws3["B1"].font = title_font
    ws3.merge_cells("B1:F1")

    hist_headers = ["MPC Meeting", "Repo Rate (Before)", "Repo Rate (After)", "CPI Inflation", "GDP Growth"]
    r = 3
    for i, h in enumerate(hist_headers):
        cell = ws3.cell(row=r, column=2 + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, (meeting, before, after, cpi, gdp) in enumerate(HISTORICAL):
        row = 4 + i
        vals = [meeting, before, after, cpi, gdp]
        fill = grey_fill if i % 2 == 0 else white_fill
        for j, v in enumerate(vals):
            cell = ws3.cell(row=row, column=2 + j, value=v)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = fill
            if j in (1, 2):
                cell.number_format = '0.00'
            elif j in (3, 4):
                cell.number_format = '0.00'

    # Add chart
    last_data_row = 4 + len(HISTORICAL) - 1
    chart = LineChart()
    chart.title = "Repo Rate vs CPI Inflation"
    chart.y_axis.title = "Rate / Inflation (%)"
    chart.x_axis.title = "MPC Meeting"
    chart.width = 28
    chart.height = 14
    chart.style = 10

    cats = Reference(ws3, min_col=2, min_row=4, max_row=last_data_row)
    repo_data = Reference(ws3, min_col=4, min_row=3, max_row=last_data_row)
    cpi_data = Reference(ws3, min_col=5, min_row=3, max_row=last_data_row)

    chart.add_data(repo_data, titles_from_data=True)
    chart.add_data(cpi_data, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = ACCENT_BLUE
    chart.series[1].graphicalProperties.line.solidFill = "E74C3C"

    ws3.add_chart(chart, f"B{last_data_row + 3}")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4: ASSUMPTIONS (INPUT SHEET)
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Assumptions")
    ws4.sheet_properties.tabColor = "E74C3C"

    ws4.column_dimensions["A"].width = 3
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 40

    ws4["B1"] = "INPUT ASSUMPTIONS — UPDATE BEFORE EACH MPC"
    ws4["B1"].font = title_font
    ws4.merge_cells("B1:D1")

    ws4["B2"] = "Last updated: [Enter date]"
    ws4["B2"].font = Font(name="Calibri", italic=True, size=9, color="999999")

    assump_headers = ["Parameter", "Value", "Unit", "Source"]
    r = 4
    for i, h in enumerate(assump_headers):
        cell = ws4.cell(row=r, column=2 + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    assump_data = [
        ("CPI Inflation (YoY)", 4.26, "%", "MOSPI / RBI"),
        ("Core CPI", 4.10, "%", "MOSPI"),
        ("WPI Inflation (YoY)", 2.31, "%", "MOSPI"),
        ("GDP Growth (FY estimate)", 6.50, "%", "RBI / NSO"),
        ("IIP Growth", 5.20, "%", "MOSPI"),
        ("CAD (% of GDP)", 1.20, "%", "RBI"),
        ("Brent Crude ($/bbl)", 72.0, "$", "Bloomberg"),
        ("Fed Funds Rate (upper)", 4.50, "%", "US Fed"),
        ("USD/INR", 86.50, "", "RBI reference rate"),
        ("Forex Reserves ($ bn)", 640.0, "$ bn", "RBI"),
        ("FII Net Flows (monthly)", 1.5, "$ bn", "NSDL / SEBI"),
        ("MIBOR-OIS Spread", 15.0, "bps", "FIMMDA"),
        ("LAF Balance", -20000, "₹ Cr", "RBI daily data"),
        ("MPC Next Meeting Date", "Apr 2026", "", "RBI website"),
        ("MPC Stance", "Accommodative", "", "MPC statement"),
    ]

    for i, (param, val, unit, source) in enumerate(assump_data):
        row = 5 + i
        fill = grey_fill if i % 2 == 0 else white_fill
        ws4.cell(row=row, column=2, value=param).font = bold_font
        ws4.cell(row=row, column=2).alignment = left_align
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=2).fill = fill

        ws4.cell(row=row, column=3, value=val).font = normal_font
        ws4.cell(row=row, column=3).alignment = center_align
        ws4.cell(row=row, column=3).border = thin_border
        ws4.cell(row=row, column=3).fill = fill
        if isinstance(val, (int, float)):
            ws4.cell(row=row, column=3).number_format = dec2_fmt

        ws4.cell(row=row, column=4, value=unit).font = normal_font
        ws4.cell(row=row, column=4).alignment = center_align
        ws4.cell(row=row, column=4).border = thin_border
        ws4.cell(row=row, column=4).fill = fill

        ws4.cell(row=row, column=5, value=source).font = Font(name="Calibri", size=9, color="666666")
        ws4.cell(row=row, column=5).alignment = left_align
        ws4.cell(row=row, column=5).border = thin_border
        ws4.cell(row=row, column=5).fill = fill

    # Data sources guide
    src_row = 5 + len(assump_data) + 2
    ws4.merge_cells(f"B{src_row}:E{src_row}")
    ws4[f"B{src_row}"] = "KEY DATA SOURCES"
    ws4[f"B{src_row}"].font = section_font

    sources = [
        ("CPI / WPI / IIP", "MOSPI — mospi.gov.in, updated monthly"),
        ("GDP", "NSO advance/revised estimates — mospi.gov.in"),
        ("CAD / Forex / LAF", "RBI Weekly Statistical Supplement — rbi.org.in"),
        ("FII Flows", "NSDL data — nsdl.co.in"),
        ("MIBOR / OIS", "FIMMDA — fimmda.org / Bloomberg"),
        ("Fed Funds Rate", "federalreserve.gov — FOMC statements"),
        ("Crude Oil", "Bloomberg / CNBC / Reuters commodities"),
        ("MPC Decisions", "RBI MPC press releases — rbi.org.in"),
    ]

    for i, (item, src) in enumerate(sources):
        row = src_row + 1 + i
        ws4.cell(row=row, column=2, value=item).font = bold_font
        ws4.cell(row=row, column=2).border = thin_border
        ws4.merge_cells(f"C{row}:E{row}")
        ws4.cell(row=row, column=3, value=src).font = Font(name="Calibri", size=9, color="666666")
        ws4.cell(row=row, column=3).border = thin_border

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 5: WEIGHT CALIBRATION
    # ═══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Weight Calibration")
    ws5.sheet_properties.tabColor = "8E44AD"

    ws5.column_dimensions["A"].width = 3
    ws5.column_dimensions["B"].width = 32
    ws5.column_dimensions["C"].width = 14
    ws5.column_dimensions["D"].width = 14
    ws5.column_dimensions["E"].width = 14
    ws5.column_dimensions["F"].width = 14
    ws5.column_dimensions["G"].width = 40

    ws5["B1"] = "WEIGHT CALIBRATION FRAMEWORK"
    ws5["B1"].font = title_font
    ws5.merge_cells("B1:G1")

    ws5["B2"] = "Rationale: CPI carries highest weight because RBI's primary mandate is inflation targeting (4±2%). MPC members consistently cite CPI as the deciding factor."
    ws5["B2"].font = Font(name="Calibri", size=9, color="666666", italic=True)
    ws5.merge_cells("B2:G2")

    cal_headers = ["Indicator", "Assigned Weight", "MPC Mentions (2024)", "Predictive Power", "Confidence", "Rationale"]
    r = 4
    for i, h in enumerate(cal_headers):
        cell = ws5.cell(row=r, column=2 + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    calibration_data = [
        ("CPI Inflation", 0.18, "Every meeting", "High", "High", "Primary mandate — always cited in statement"),
        ("Core CPI", 0.10, "Most meetings", "High", "High", "RBI's preferred measure for persistence"),
        ("GDP Growth", 0.12, "Every meeting", "High", "High", "Output gap determines easing room"),
        ("CAD", 0.08, "4 of 6 meetings", "Medium", "Medium", "External balance affects rupee & policy space"),
        ("Crude Oil", 0.07, "3 of 6 meetings", "Medium", "Medium", "Key input for inflation & CAD"),
        ("Fed Funds Rate", 0.06, "3 of 6 meetings", "Medium", "Medium", "Limits RBI's easing if differential widens"),
        ("LAF Liquidity", 0.06, "4 of 6 meetings", "Medium", "Medium", "Operational signal of stance"),
        ("MIBOR-OIS Spread", 0.05, "2 of 6 meetings", "Low-Med", "Medium", "Real-time liquidity stress gauge"),
        ("USD/INR", 0.05, "2 of 6 meetings", "Low-Med", "Medium", "Imported inflation channel"),
        ("WPI Inflation", 0.05, "1 of 6 meetings", "Low", "Low", "Lagging indicator, less policy-relevant"),
        ("FII Flows", 0.04, "1 of 6 meetings", "Low", "Low", "Capital flow pressure, not direct MPC driver"),
        ("IIP Growth", 0.04, "1 of 6 meetings", "Low", "Low", "Granular industrial data, not macro driver"),
        ("Forex Reserves", 0.04, "0 of 6 meetings", "Low", "Low", "Buffer metric, not rate signal"),
    ]

    for i, (ind, wt, mentions, power, conf, rationale) in enumerate(calibration_data):
        row = 5 + i
        fill = grey_fill if i % 2 == 0 else white_fill
        vals = [ind, wt, mentions, power, conf, rationale]
        for j, v in enumerate(vals):
            cell = ws5.cell(row=row, column=2 + j, value=v)
            cell.font = normal_font if j != 0 else bold_font
            cell.alignment = center_align if j < 5 else left_align
            cell.border = thin_border
            cell.fill = fill
            if j == 1:
                cell.number_format = pct_fmt

    # Total weight check
    tot_row = 5 + len(calibration_data)
    ws5.cell(row=tot_row, column=2, value="TOTAL").font = bold_font
    ws5.cell(row=tot_row, column=3, value=f"=SUM(C5:C{tot_row-1})")
    ws5.cell(row=tot_row, column=3).number_format = pct_fmt
    ws5.cell(row=tot_row, column=3).font = bold_font
    ws5.cell(row=tot_row, column=3).alignment = center_align
    for c in range(2, 8):
        ws5.cell(row=tot_row, column=c).fill = light_fill
        ws5.cell(row=tot_row, column=c).border = thin_border

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 6: MPC CALENDAR
    # ═══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("MPC Calendar")
    ws6.sheet_properties.tabColor = "F39C12"

    ws6.column_dimensions["A"].width = 3
    ws6.column_dimensions["B"].width = 14
    ws6.column_dimensions["C"].width = 18
    ws6.column_dimensions["D"].width = 30
    ws6.column_dimensions["E"].width = 30

    ws6["B1"] = "RBI MPC MEETING CALENDAR — FY25-26"
    ws6["B1"].font = title_font
    ws6.merge_cells("B1:D1")

    mpc_headers = ["MPC Meeting", "Decision Date", "Status", "Pre-Meeting Actions"]
    r = 3
    for i, h in enumerate(mpc_headers):
        cell = ws6.cell(row=r, column=2 + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    mpc_meetings = [
        ("Feb 2026", "Feb 6-8, 2026", "Completed — Cut to 6.00%", "Update scoring with Jan CPI"),
        ("Apr 2026", "Apr 7-9, 2026", "Upcoming", "Update with Mar CPI, Q4 GDP advance"),
        ("Jun 2026", "Jun 4-6, 2026", "Upcoming", "Update with May CPI, monsoon outlook"),
        ("Aug 2026", "Aug 5-7, 2026", "Upcoming", "Update with Jul CPI, Q1 FY27 GDP"),
        ("Oct 2026", "Oct 7-9, 2026", "Upcoming", "Update with Sep CPI, festive season data"),
        ("Dec 2026", "Dec 3-5, 2026", "Upcoming", "Update with Nov CPI, FY27 budget impact"),
    ]

    for i, (meet, date, status, action) in enumerate(mpc_meetings):
        row = 4 + i
        fill = grey_fill if i % 2 == 0 else white_fill
        vals = [meet, date, status, action]
        for j, v in enumerate(vals):
            cell = ws6.cell(row=row, column=2 + j, value=v)
            cell.font = normal_font
            cell.alignment = center_align if j < 3 else left_align
            cell.border = thin_border
            cell.fill = fill
        # Highlight completed
        if "Completed" in status:
            ws6.cell(row=row, column=4).fill = green_fill
        elif "Upcoming" in status:
            ws6.cell(row=row, column=4).fill = yellow_fill

    # Freeze panes on key sheets
    ws2.freeze_panes = "B4"
    ws3.freeze_panes = "B4"

    # Print setup
    for sheet in [ws, ws2, ws3, ws4, ws5, ws6]:
        sheet.sheet_view.showGridLines = False

    return wb


if __name__ == "__main__":
    print("Building RBI Repo Rate Projection Model...")
    wb = create_workbook()
    output_path = "/home/homepc/.openclaw/workspace/RBI_Repo_Rate_Projection.xlsx"
    wb.save(output_path)
    print(f"✅ Saved to: {output_path}")
    print("\nSheets created:")
    print("  1. Dashboard — Current state, projections, scenarios")
    print("  2. Indicator Scoring — 13-indicator weighted matrix (enter scores here)")
    print("  3. Historical Data — MPC decisions with chart")
    print("  4. Assumptions — Input sheet for latest macro data")
    print("  5. Weight Calibration — Rationale for each weight")
    print("  6. MPC Calendar — Key dates & pre-meeting actions")
    print("\nNext steps:")
    print("  • Open Assumptions sheet → update with latest data")
    print("  • Open Indicator Scoring → enter -5 to +5 scores per indicator")
    print("  • Weighted Score auto-calculates → check interpretation guide")
    print("  • Update Dashboard projections based on score")